"""
Módulo de Consolidação de Dados
Gera planilhas Excel com os dados extraídos e validados
"""

import pandas as pd
from typing import List, Dict
from decimal import Decimal
from extractor import NFSeData
from validator import NFSeValidator, ValidationIssue
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class NFSeConsolidator:
    """Consolidador de dados de NFS-e em planilhas"""
    
    def __init__(self):
        self.validator = NFSeValidator()
    
    def consolidate_to_excel(
        self, 
        data_list: List[NFSeData], 
        output_path: str,
        include_validation: bool = True
    ) -> str:
        """
        Consolida dados em planilha Excel
        
        Args:
            data_list: Lista de dados de NFS-e
            output_path: Caminho para salvar o arquivo Excel
            include_validation: Se deve incluir planilha de validação
            
        Returns:
            Caminho do arquivo gerado
        """
        # Cria DataFrame principal
        df_main = self._create_main_dataframe(data_list)
        
        # Cria DataFrame de validação se solicitado
        validation_data = []
        if include_validation:
            for data in data_list:
                issues = self.validator.validate(data)
                validation_data.append({
                    'Número NF': data.numero_nota,
                    'Arquivo': data.arquivo_origem,
                    'Status': self.validator.get_validation_status(issues),
                    'Quantidade Problemas': len(issues),
                    'Erros': len([i for i in issues if i.severity == 'ERROR']),
                    'Avisos': len([i for i in issues if i.severity == 'WARNING']),
                    'Detalhes': self.validator.generate_validation_report(issues)
                })
        
        df_validation = pd.DataFrame(validation_data) if validation_data else None
        
        # Salva Excel com formatação
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_main.to_excel(writer, sheet_name='Dados NFS-e', index=False)
            
            if df_validation is not None:
                df_validation.to_excel(writer, sheet_name='Validação', index=False)
            
            # Cria planilha resumo
            df_summary = self._create_summary_dataframe(data_list)
            df_summary.to_excel(writer, sheet_name='Resumo', index=False)
        
        # Aplica formatação
        self._apply_formatting(output_path)
        
        return output_path
    
    def _create_main_dataframe(self, data_list: List[NFSeData]) -> pd.DataFrame:
        """Cria DataFrame principal com todos os dados"""
        rows = []
        
        for data in data_list:
            row = {
                # Identificação
                'Número NF': data.numero_nota,
                'Data Emissão': data.data_emissao,
                'Data Competência': data.data_competencia,
                'Município': data.municipio,
                'Chave Acesso': data.chave_acesso,
                
                # Prestador
                'Prestador CNPJ': data.prestador_cnpj,
                'Prestador Nome': data.prestador_nome,
                'Prestador IM': data.prestador_inscricao,
                
                # Tomador
                'Tomador CNPJ': data.tomador_cnpj,
                'Tomador Nome': data.tomador_nome,
                'Tomador IM': data.tomador_inscricao,
                
                # Valores principais
                'Valor Serviços': float(data.valor_servicos),
                'Desconto': float(data.valor_desconto),
                'Base Cálculo': float(data.base_calculo),
                
                # ISS
                'ISS Alíquota (%)': float(data.iss_aliquota),
                'ISS Valor': float(data.iss_valor),
                'ISS Retido': data.iss_retido,
                'Município Retenção': data.municipio_retencao,
                
                # Tributos Federais
                'PIS': float(data.pis),
                'COFINS': float(data.cofins),
                'CSLL': float(data.csll),
                'IRRF': float(data.irrf),
                'INSS': float(data.inss),
                'Outras Retenções': float(data.outras_retencoes),
                
                # Totais
                'Total Tributos Retidos': float(data.total_tributos_retidos),
                'Valor Líquido': float(data.valor_liquido),
                
                # Informações adicionais
                'Descrição Serviço': data.descricao_servico,
                'Código Atividade': data.codigo_atividade,
                'Arquivo Origem': data.arquivo_origem,
                'Padrão Layout': data.padrao_layout
            }
            
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _create_summary_dataframe(self, data_list: List[NFSeData]) -> pd.DataFrame:
        """Cria DataFrame com resumo estatístico"""
        total_notas = len(data_list)
        
        # Calcula totalizadores
        total_servicos = sum(float(d.valor_servicos) for d in data_list)
        total_iss = sum(float(d.iss_valor) for d in data_list)
        total_pis = sum(float(d.pis) for d in data_list)
        total_cofins = sum(float(d.cofins) for d in data_list)
        total_csll = sum(float(d.csll) for d in data_list)
        total_irrf = sum(float(d.irrf) for d in data_list)
        total_inss = sum(float(d.inss) for d in data_list)
        total_outras = sum(float(d.outras_retencoes) for d in data_list)
        total_tributos = sum(float(d.total_tributos_retidos) for d in data_list)
        total_liquido = sum(float(d.valor_liquido) for d in data_list)
        
        # Conta notas com ISS retido
        notas_iss_retido = len([d for d in data_list if d.iss_retido == "Sim"])
        
        # Agrupa por município
        municipios = {}
        for data in data_list:
            mun = data.municipio
            if mun not in municipios:
                municipios[mun] = {'count': 0, 'valor': 0}
            municipios[mun]['count'] += 1
            municipios[mun]['valor'] += float(data.valor_servicos)
        
        # Agrupa por prestador
        prestadores = {}
        for data in data_list:
            prest = data.prestador_nome
            if prest not in prestadores:
                prestadores[prest] = {'count': 0, 'valor': 0}
            prestadores[prest]['count'] += 1
            prestadores[prest]['valor'] += float(data.valor_servicos)
        
        summary_data = [
            {'Métrica': 'TOTAIS GERAIS', 'Valor': ''},
            {'Métrica': 'Total de Notas Fiscais', 'Valor': total_notas},
            {'Métrica': 'Valor Total dos Serviços', 'Valor': f'R$ {total_servicos:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')},
            {'Métrica': 'Valor Total Líquido', 'Valor': f'R$ {total_liquido:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')},
            {'Métrica': '', 'Valor': ''},
            
            {'Métrica': 'TRIBUTOS RETIDOS', 'Valor': ''},
            {'Métrica': 'ISS Total', 'Valor': f'R$ {total_iss:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')},
            {'Métrica': 'PIS Total', 'Valor': f'R$ {total_pis:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')},
            {'Métrica': 'COFINS Total', 'Valor': f'R$ {total_cofins:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')},
            {'Métrica': 'CSLL Total', 'Valor': f'R$ {total_csll:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')},
            {'Métrica': 'IRRF Total', 'Valor': f'R$ {total_irrf:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')},
            {'Métrica': 'INSS Total', 'Valor': f'R$ {total_inss:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')},
            {'Métrica': 'Outras Retenções', 'Valor': f'R$ {total_outras:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')},
            {'Métrica': 'Total Geral de Tributos', 'Valor': f'R$ {total_tributos:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')},
            {'Métrica': '', 'Valor': ''},
            
            {'Métrica': 'RETENÇÕES', 'Valor': ''},
            {'Métrica': 'Notas com ISS Retido', 'Valor': f'{notas_iss_retido} ({notas_iss_retido/total_notas*100:.1f}%)'},
            {'Métrica': 'Notas sem ISS Retido', 'Valor': f'{total_notas - notas_iss_retido} ({(total_notas-notas_iss_retido)/total_notas*100:.1f}%)'},
            {'Métrica': '', 'Valor': ''},
            
            {'Métrica': 'POR MUNICÍPIO', 'Valor': ''},
        ]
        
        for mun, info in sorted(municipios.items(), key=lambda x: x[1]['valor'], reverse=True):
            summary_data.append({
                'Métrica': f'  {mun}',
                'Valor': f'{info["count"]} notas - R$ {info["valor"]:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            })
        
        summary_data.append({'Métrica': '', 'Valor': ''})
        summary_data.append({'Métrica': 'TOP 5 PRESTADORES', 'Valor': ''})
        
        for prest, info in sorted(prestadores.items(), key=lambda x: x[1]['valor'], reverse=True)[:5]:
            summary_data.append({
                'Métrica': f'  {prest}',
                'Valor': f'{info["count"]} notas - R$ {info["valor"]:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            })
        
        return pd.DataFrame(summary_data)
    
    def _apply_formatting(self, file_path: str):
        """Aplica formatação visual ao Excel"""
        wb = load_workbook(file_path)
        
        # Formatação da planilha principal
        if 'Dados NFS-e' in wb.sheetnames:
            ws = wb['Dados NFS-e']
            self._format_sheet(ws, is_data=True)
        
        # Formatação da planilha de validação
        if 'Validação' in wb.sheetnames:
            ws = wb['Validação']
            self._format_sheet(ws, is_validation=True)
        
        # Formatação da planilha de resumo
        if 'Resumo' in wb.sheetnames:
            ws = wb['Resumo']
            self._format_sheet(ws, is_summary=True)
        
        wb.save(file_path)
    
    def _format_sheet(self, ws, is_data=False, is_validation=False, is_summary=False):
        """Aplica formatação a uma planilha"""
        # Define estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formata cabeçalho
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Ajusta largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Aplica bordas e alinhamento
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                
                # Alinha valores numéricos à direita
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Congela primeira linha
        ws.freeze_panes = 'A2'
        
        # Formatação específica para resumo
        if is_summary:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[0].value and any(x in str(row[0].value) for x in ['TOTAIS', 'TRIBUTOS', 'RETENÇÕES', 'MUNICÍPIO', 'PRESTADORES']):
                    row[0].font = Font(bold=True, size=11)
                    row[0].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
